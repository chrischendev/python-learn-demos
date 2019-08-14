import openpyxl

# 创建一个工作簿
workbook = openpyxl.Workbook()
workbook.encoding = 'UTF-8'

# 创建工作表
sheet = workbook.create_sheet()
sheet.title = 'first_sheet'

# 写入单元格数据
sheet.cell(1, 1, 'ID')
sheet.cell(1, 2, 'Title')
sheet.cell(1, 3, 'Code')
sheet.cell(1, 4, 'Url')

# 批量写入
infos = [1, '乱港分子都去死', 'MMNHGGFC74523956309', 'http://ashfqowehjs.vcbjmAJJJ.CB']
for i in range(len(infos)):
    sheet.cell(2, i + 1, infos[i])

# 保存文件
workbook.save(filename='D:\\python_test_xlsx_write.xlsx')
