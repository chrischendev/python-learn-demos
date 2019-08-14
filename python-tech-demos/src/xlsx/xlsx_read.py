import openpyxl

# 获取工作簿对象
workbook = openpyxl.load_workbook("D:\\python_test_xlsx.xlsx")

# 获取工作簿 workbook的所有工作表
sheet_names = workbook.sheetnames
print(sheet_names)
print(sheet_names[0])
print(workbook[sheet_names[0]])

# 获取工作表对象
sheet1 = workbook['Sheet1']
print(sheet1)
print(sheet1.title)
sheet2 = workbook.worksheets[0]
print(sheet2)

# 获取当前活动的表
sheet3 = workbook.active
print(sheet3)

# 读取单元格
print(sheet1.cell(2, 2).value)

for row in sheet1.rows:
    for cell in row:
        print(cell.value, end='\t')
    print('')

# 输出一行数据 第二行
for cell in list(sheet1.rows)[1]:
    print(cell.value)
